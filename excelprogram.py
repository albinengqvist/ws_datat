import openpyxl
import os


filename = 'G:/Produktion/data/ws_data.xlsx'

if os.path.isfile(filename):
   wb = openpyxl.Workbook(filename)
   if 'Data' in wb.sheetnames:
       pass
   else:
       wb.create_sheet(index=0, title='Data')
else:
   wb = openpyxl.Workbook()
   wb.create_sheet(index=0, title='Data')
   wb.save(filename)

wb = openpyxl.load_workbook(filename)
sheet = wb['Data']

def add_wsdata(order_no, art_no, art_fam, open, closing, comment):
    ws = wb.active
    first_column = ws['A']
    second_column = ws['B']
    third_column = ws['C']
    fourth_column = ws['D']
    fifth_column = ws['E']
    sixth_column = ws['F']

    col_len1 = str(len(first_column)+1)
    col_len2 = str(len(second_column)+1)
    col_len3 = str(len(third_column)+1)
    col_len4 = str(len(fourth_column)+1)
    col_len5 = str(len(fifth_column)+1)
    col_len5 = str(len(sixth_column)+1)

    sheet['A' + col_len1] = order_no
    sheet['B' + col_len2] = art_no
    sheet['C' + col_len2] = art_fam
    sheet['D' + col_len2] = open
    sheet['E' + col_len2] = closing
    sheet['F' + col_len2] = comment
    wb.save(filename)

if (sheet['A1'].value == 'order_no') and (sheet['B1'].value == 'art_no'):
    pass
else:
    sheet['A1'] = 'order_no'
    sheet['B1'] = 'art_no'
    sheet['C1'] = 'art_fam'
    sheet['D1'] = 'open'
    sheet['F1'] = 'comment'

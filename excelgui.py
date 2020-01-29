import openpyxl
import os
from tkinter import *
from excelprogram import *


root = Tk()


def cont():
    entry3.delete(0, END)
    entry4.delete(0, END)
    entry4.delete(0, END)


def clear():
    entry1.delete(0, END)
    entry2.delete(0, END)
    entry3.delete(0, END)
    entry4.delete(0, END)
    entry5.delete(0, END)


def popup():
    win = Toplevel()
    win.wm_title("Wapro")
    win.minsize(200,100)
    win.resizable(width=False, height=False)
    win.grid_columnconfigure(0, minsize=20)
    win.grid_columnconfigure(2, minsize=5)
    win.grid_columnconfigure(4, minsize=5)
    win.grid_columnconfigure(6, minsize=20)
    win.grid_rowconfigure(0, minsize=20)
    win.grid_rowconfigure(2, minsize=20)
    win.grid_rowconfigure(4, minsize=10)

    l = Label(win, text="Sparat!", font = (None, 15))
    l.grid(row=1, column=1)

    b= Button(win, text="Fortsätt samma order", command=lambda:[cont(),win.destroy()])
    b.grid(row=3, column=1)
    c = Button(win, text="Ny order", command=lambda:[clear(),win.destroy()])
    c.grid(row=3, column=3)
    d = Button(win, text="Stäng", command=lambda:[win.destroy(),root.destroy()])
    d.grid(row=3, column=5)

def saveData():
    global entry1
    global entry2
    global entry3
    global entry4
    global entry5
    global entry6

    order_no = entry1.get()
    art_no = entry2.get()
    open = entry3.get()
    closing = entry4.get()
    comment = entry5.get()

    try:
        art_fam = (art_no.split('-')[0] + '-' + art_no.split('-')[1])
    except:
        art_fam = art_no

    add_wsdata(order_no, art_no, art_fam, open, closing, comment)

filename = 'G:/Produktion/data/ws_data_hist.xlsx'

wb = openpyxl.load_workbook(filename)
ws = wb['data']

def check_op(sv):
    sv.get
    global entry2
    global min_op
    global max_op
    global min_cp
    global max_cp

    art_no = entry2.get()
    try:
        art_fam = (art_no.split('-')[0] + '-' + art_no.split('-')[1])
    except:
        art_fam = art_no
    min_op = '000'
    max_op = '000'
    min_cp = '000'
    max_cp = '000'
    #print('art_fam:', art_fam)
    row_count = 0

    for cell in ws['A']:
        row_count = row_count +1
        try:
            if str(art_fam).lower() == str(cell.value).lower():
                #print('success')

                min_op = int(ws.cell(row=cell.row, column= 3).value)*0.85
                max_op = int(ws.cell(row=cell.row, column= 3).value)*1.15
                min_cp = int(ws.cell(row=cell.row, column= 2).value)*0.85
                max_cp = int(ws.cell(row=cell.row, column= 2).value)*1.15
                #print(max_cp)

        except:
            pass


    return min_op, max_op, min_cp, max_cp

root.minsize(400,300)
root.title("WaStop Mätdata")
root.resizable(0, 0)
root.grid_columnconfigure(1, minsize=30)
root.grid_columnconfigure(5, minsize=30)
root.grid_rowconfigure(0, minsize=50)
root.grid_rowconfigure(2, minsize=50)
root.grid_rowconfigure(4, minsize=50)
root.grid_rowconfigure(6, minsize=50)
root.grid_rowconfigure(8, minsize=20)



label_1 = Label(root,text = "Ordernummer:", font = (None, 11))
label_1.grid(row=1,column=2,sticky=W)

entry1 = Entry(root, width = 30, relief=RIDGE)
entry1.grid(row=1, column=3, sticky=W)
sv= StringVar()
sv.trace('w', lambda name, index, mode, sv=sv: callback(sv))
label_2 = Label(root, text = "Artikelnummer:", font = (None, 11))
label_2.grid(row=2,column=2,sticky=W)

entry2 = Entry(root, width = 30, textvariable=sv, relief=RIDGE)
entry2.grid(row=2, column=3, sticky=W)

label_3 = Label(root,text = "Öppningstryck [mmH2O]:", font = (None, 11))
label_3.grid(row=3,column=2,sticky=W)
min_op, max_op, min_cp, max_cp = check_op(sv)
text_var1 = StringVar()
label_3_1 = Label(root, textvariable=text_var1, font = (None, 11))
label_3_1.grid(row=3, column=4, sticky=E)
def callback(sv):
    min_op, max_op, min_cp, max_cp = check_op(sv)
    str1 = (str((round(int(min_op), 0)))+ " - "+str((round(int(max_op), 0))))
    str2 = (str((round(int(min_cp), 0)))+ " - "+str((round(int(max_cp), 0))))
    text_var1.set(str1)
    text_var2.set(str2)

entry3 = Entry(root, width = 30, relief=RIDGE)
entry3.grid(row=3, column=3, sticky=W)

label_4 = Label(root,text = "Stängningstryck [mmH2O]:", font = (None, 11))
label_4.grid(row=4,column=2,sticky=W)
text_var2 = StringVar()
label_4_1 = Label(root, textvariable=text_var2, font = (None, 11))
label_4_1.grid(row=4, column=4, sticky=E)

entry4 = Entry(root, width = 30, relief=RIDGE)
entry4.grid(row=4, column=3, sticky=W)

label_5 = Label(root,text = "Kommentar:", font = (None, 11))
label_5.grid(row=5,column=2,sticky=W)

entry5 = Entry(root, width = 30, relief=RIDGE)
entry5.grid(row=5, column=3, sticky=W)

button = Button(root,text='Spara',command=lambda:[saveData(),popup()], height = 1, width = 10,font = (None, 11))
button.grid(row=7, column=3, sticky=E)
button.bind("<Return>", (lambda event: [saveData(),popup()]))



background_image=PhotoImage(file = 'G:/Produktion/data/wapro.gif')
background_label = Label(root, image=background_image)
background_label.grid(row=7, column=2, sticky=W)

root.mainloop()
